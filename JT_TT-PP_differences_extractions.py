#------------------------------------------------------------------------------------
# Name:        Comparing Two Datasets (Excel Sheets), Measuring Trial Lengths
# Purpose:     Takes input from an excel file, finds lines in a certain column
#              with multiple lines in them, then spreads it out into multiple lines
#              Also searches ClinicalTrials.gov for the NCT#'s of the bigger dataset.
#
# Author:      Jonathan Thomas
#
# Copyright:   (c) WOAH 2018
# License:     MIT License
#------------------------------------------------------------------------------------

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from collections import OrderedDict
import StringIO
import re
import unicodedata

import Tkinter
import tkFileDialog
import html5lib
import time
import random


#remember to have requests, BeautifulSoup4, requests[security], and in your packages for this to work!
import requests
from bs4 import BeautifulSoup
#Define starting constants:

#Beginning and Ending Cells in your worksheet to Process
StartCell = 2
EndCell = 4632
EndCell2 = 3323

def search(text,n):
    '''Searches for text, and retrieves n words either side of the text, which are returned seperately'''
    word = r"\W*([\w]+)"
    groups = re.search(r'{}\W*{}{}'.format(word*n,'place',word*n), text).groups()
    return groups[:n],groups[n:]

def find_all(a_str, sub):
    start = 0
    while True:
        start = a_str.find(sub, start)
        if start == -1: return
        yield start
        start += len(sub) # use start += 1 to find overlapping matches


def find_between_r( s, first, last ):
    try:
        start = s.rindex( first ) + len( first )
        end = s.rindex( last, start )
        return s[start:end]
    except ValueError:
        return ""

#Open the file in question:

open_file = tkFileDialog.askopenfilename ()

#Dialog Box opening complete

print("Now opening: " + str(open_file))

#Performing initial IO on workbooks
wb = load_workbook(filename = open_file)
blank = Workbook()
ws = wb.active

sheet_ranges = wb['Drugs (The Original)']
second_sheet_ranges = wb['trialtrove_4855441']
Input_Data = wb['Results']
FinishedLinesTest = []
alreadyChecked = []
InputRowCounter = StartCell

p1NCTS = []
p2NCTS = []
p3NCTS = []
p1ncts = []
p2ncts = []
p3ncts = []

TTncts = []
ttncts = []
ppncts = []

Overlaps = []
#Find the NCT#'s of the Phase 3 Trials from Pharmaprojects Data
while(InputRowCounter != EndCell):   
    print(InputRowCounter)
    
    OBJ1 = sheet_ranges['BC' + str(InputRowCounter)].value
    print(OBJ1)
    if(OBJ1 is None):
        InputRowCounter += 1
        continue

    PCSubjects = re.findall(r"NCT\d\d\d\d\d\d\d\d", OBJ1)
    if(PCSubjects == [] or PCSubjects == ['-']):
        InputRowCounter += 1
        continue
    
    i = 0
    while(i<len(PCSubjects)):
        PCSubjects[i] = PCSubjects[i].encode('utf-8')
        i += 1
        
    p1NCTS.extend(PCSubjects)
    
    InputRowCounter += 1

if(InputRowCounter == EndCell):
    p1ncts = list(set(p1NCTS))

#Find the NCT#'s of the Phase 2 Trials from Pharmaprojects Data
InputRowCounter = 2
while(InputRowCounter != EndCell):   

    OBJ2 = sheet_ranges['BD' + str(InputRowCounter)].value
    if(OBJ2 is None):
        InputRowCounter += 1
        continue
    
    PCSubjects = re.findall(r"NCT\d\d\d\d\d\d\d\d", OBJ2)
    if(PCSubjects == [] or PCSubjects == ['-']):
        InputRowCounter += 1
        continue
    
    i = 0
    while(i<len(PCSubjects)):
        PCSubjects[i] = PCSubjects[i].encode('utf-8')
        i += 1
        
    p2NCTS.extend(PCSubjects)
    
    InputRowCounter += 1

if(InputRowCounter==EndCell):
    p2ncts = list(set(p2NCTS))

#Find the NCT#'s of the Phase 1 Trials from Pharmaprojects Data
InputRowCounter = 2    
while(InputRowCounter != EndCell):    

    OBJ3 = sheet_ranges['BE' + str(InputRowCounter)].value
    if(OBJ3 is None):
        InputRowCounter += 1
        continue    
    PCSubjects = re.findall(r"NCT\d\d\d\d\d\d\d\d", OBJ3)
    if(PCSubjects == [] or PCSubjects == ['-']):
        InputRowCounter += 1
        continue
    
    i = 0
    while(i<len(PCSubjects)):
        PCSubjects[i] = PCSubjects[i].encode('utf-8')
        i += 1
        
    p3NCTS.extend(PCSubjects)

    InputRowCounter += 1    

if(InputRowCounter == EndCell):
    p3ncts = list(set(p3NCTS))

#Find the NCT#'s of the 'Trial Identifier' column from TrialTrove Data
InputRowCounter = 2
while(InputRowCounter != EndCell2):    

    OBJ4 = second_sheet_ranges['J' + str(InputRowCounter)].value
    if(OBJ4 is None):
        InputRowCounter += 1
        continue
    
    PCSubjects = re.findall(r"NCT\d\d\d\d\d\d\d\d", OBJ4)
    if(PCSubjects == [] or PCSubjects == ['-']):
        InputRowCounter += 1
        continue
    
    i = 0
    while(i<len(PCSubjects)):
        PCSubjects[i] = PCSubjects[i].encode('utf-8')
        i += 1
        
    TTncts.extend(PCSubjects)

    InputRowCounter += 1    

#Generate TrialTrove List of NCT#'s...without duplicates
if(InputRowCounter == EndCell2):
    ttncts = list(set(TTncts))

ppncts.extend(p1ncts)
ppncts.extend(p2ncts)
ppncts.extend(p3ncts)
#Generate Pharmaprojects List of NCT#'s...without duplicates
ppncts = list(set(ppncts))

Overlaps = (set(ttncts) & set(ppncts))
print('There are ' + str(len(Overlaps)) + ' similar NCT Numbers.')
if(set(ttncts) == set(ppncts)):
    print("The data is the same on both TT & PP :) ")
else:
    print("Use whichever one is bigger.")   
    
if(len(ttncts) == len(ppncts)):
    print("But they are the same size")
elif(len(ttncts) > len(ppncts)):
    print("Use TrialTrove!")
elif(len(ttncts) < len(ppncts)):
    print("Use PharmaProjects!")
else:
    print("....")   
    
differences = [x for x in ttncts if x not in ppncts] + [x for x in ppncts if x not in ttncts]
print('There are ' + str(len(differences)) + ' different NCT Numbers.')

#Searching Clinicaltrials.gov using NCT#'s to find data, specifically accurate and reliable start and completion dates.
ResultsRowCounter = 2       
h = 1
i = 1
while(i<len(differences)):
    #Saving the excel file after every fifteen search attempts. 
    if(i%15 == 0):
        wb.save(open_file)
        
    #Putting the process to sleep, to help prevent potential connection issues
    secondBotTrick = 0
    secondBotTrickSec = 0
    secondBotTrickSec = random.random() * random.choice([5, 10])
    secondBotTrick = int(secondBotTrickSec)
    time.sleep(secondBotTrick)

    #Find Queries and count number
    url = 'https://clinicaltrials.gov/ct2/results?cond=&term=' + differences[h] + '&cntry=&state=&city=&dist='
    r=requests.get(url)
    html = r.text
    soup = BeautifulSoup(r.text,"html5lib")

    #print(url)
    WebLines =  StringIO.StringIO(str(soup))
    FinWebLines = WebLines.readlines()

    j = 0
    DataLine = 0
    while(j<len(FinWebLines)):
        if(FinWebLines[j].find("found for:") != -1):
            #print(FinWebLines[j])
            DataLine = j
            break
        j += 1

    DataLine = FinWebLines[j].split()
    j = 0
    a = 0
    while(j <len(DataLine)):
        if(DataLine[j].find("found") != -1):
            a = j-2
            break
        j += 1

    ResultsNumber = DataLine[a].split(">")[-1]
    if(ResultsNumber == "No"):
        ResultsNumber = 0

    ResultsNumber = int(ResultsNumber)

    #Get links to search results
    links = []

    for item in soup.find_all('a'):
        links.append((item.get('href')))

    links = [str(links[x]) for x in range(len(links))]
    #print(len(links))
    glinks = []
    j = 0
    while(j<len(links)):
        if((links[j].find("/ct2/show/") != -1 )):
            glinks.append("https://clinicaltrials.gov" + links[j])
        j += 1

    links = glinks
    #print(links)

    #Go into each link and check for criterion
    j = 0
    GoodStudies = []
    INCLUDE = 1
    while(j<len(links)):
        url = links[j]
        r=requests.get(url)
        html = r.text
        soup = BeautifulSoup(html,"html5lib")
        [s.extract() for s in soup(['style', 'script', '[document]', 'head', 'title'])]
        SearchText = soup.getText()
        SearchText = SearchText.encode('utf-8')
        GoodStudies.append(links[j])
        #We want ALL studies. 
        j += 1
        print("Study " + str(j) + " out of: " + str(len(links)))

    #print(GoodStudies)
    l = 0
    link = []
    ConditionWebsite = ''
    Phases = ''
    Ages = ''
    Identifier = ''
    Criteria = ''
    Summary = ''
    DetailedDescription = ''
    StartDate = ''
    CompletionDate = ''
    thirdBotTrick = 0
    thirdBotTrickSec = 0
    while(l<len(GoodStudies)):

        #Putting the process to sleep, to help prevent potential connection issues
        thirdBotTrickSec = random.random() * random.choice([15, 30, 45, 60, 75, 90])
        thirdBotTrick = int(thirdBotTrickSec)
        time.sleep(thirdBotTrick)

        url = GoodStudies[l].replace("show/","show/record/")
        r=requests.get(url)
        html = r.text
        soup = BeautifulSoup(html,"html5lib")

        tables = []
        [s.extract() for s in soup(['style', 'script', '[document]', 'head', 'title'])]
        tables = soup.getText()
        tables = tables.encode('utf-8')

        link.append(url)
        mina = list(find_all(tables,"Condition"))
        minb = list(find_all(tables,"Intervention"))
        if(minb == []):
            minb.append(0)
        if(mina == []):
            mina.append(0)

        checker = 0
        while(checker<len(minb)):
            if(minb[checker]<=min(mina) or minb[checker]>=25000):
                minb[checker] = 100000000
            checker = checker + 1
        smallcont = []
        storagecont = []
        smallstorage = []
        u = 0
        p = 0
        while(u<len(mina)):
            p = 0
            while((p<len(minb))):
                if(minb[p] > mina[u]):
                    storagecont.append(abs(mina[u] -  minb[p]))
                else:
                    storagecont.append(10000000)
                p = p + 1
            smallstorage.append(storagecont.index(min(storagecont)))
            smallcont.append(storagecont[storagecont.index(min(storagecont))])
            storagecont = []
            u = u + 1

        PosOne = mina[smallcont.index(min(smallcont))]
        PosTwo = minb[smallstorage[smallcont.index(min(smallcont))]]

        ConditionTemp = str(tables[PosOne:PosTwo]).replace("Condition","").replace("ICMJE","")
        Indication = ConditionTemp.strip(' ')
        ConditionWebsite = Indication
                    
        mina = list(find_all(tables,"Start"))
        if(mina == []):
            mina.append(0)
        minb = list(find_all(tables,"Primary"))
        if(minb == []):
            minb.append(0)

        #print(mina)
        #print(minb)
        smallcont = []
        storagecont = []
        smallstorage = []
        u = 0
        p = 0
        while(u<len(mina)):
            p = 0
            while((p<len(minb))):
                storagecont.append(abs(mina[u] -  minb[p]))
                p = p + 1
            smallstorage.append(storagecont.index(min(storagecont)))
            smallcont.append(storagecont[storagecont.index(min(storagecont))])
            storagecont = []
            u = u + 1
        PosOne = mina[smallcont.index(min(smallcont))]
        PosTwo = minb[smallstorage[smallcont.index(min(smallcont))]]
        StartDateStatus = str(tables[PosOne:PosTwo]).replace("Start","").replace("Date","").replace("ICMJE","").replace("Estimated","")
        #print(StartDateStatus)
        StartDates = re.findall(r'\w{3,8}', StartDateStatus)
        if(StartDates != []):    
            StartDateStatus = StartDateStatus.partition(StartDates[0])
            StartDateStatus = ''.join(StartDateStatus[1:3])
            StartDateStatus = StartDateStatus.strip()
            StartDate = StartDateStatus
        else:
            StartDate = 'Not Available'
            #print(StartDateStatus)

        mina = list(find_all(tables,"Completion Date"))
        if(mina == []):
            mina.append(0)
        minb = list(find_all(tables,"Primary"))
        if(minb == []):
            minb.append(0)

        checker = 0
        while(checker<len(minb)):
            if(minb[checker]<=min(mina) or minb[checker]>=8000):
                minb[checker] = 100000000
            checker = checker + 1
        #print(mina)
        #print(minb)
        smallcont = []
        storagecont = []
        smallstorage = []
        u = 0
        p = 0
        while(u<len(mina)):
            p = 0
            while((p<len(minb))):
                storagecont.append(abs(mina[u] -  minb[p]))
                p = p + 1
            smallstorage.append(storagecont.index(min(storagecont)))
            smallcont.append(storagecont[storagecont.index(min(storagecont))])
            storagecont = []
            u = u + 1
        PosOne = mina[smallcont.index(min(smallcont))]
        PosTwo = minb[smallstorage[smallcont.index(min(smallcont))]]
        
        CompDateStatus = str(tables[PosOne:PosTwo]).replace("Primary","").replace("Date","").replace("Current","").replace("Estimated","").replace("Completion","")
        CDStatus = re.findall(r'\d\d\d\d', CompDateStatus)
        if(CDStatus != []):    
            CDStatus = CompDateStatus.partition(CDStatus[0])
            CDStatus = ''.join(CDStatus[0:2])
            CDStatus = CDStatus.strip()
            CompletionDate = CDStatus
        else:
            CompletionDate = 'Not Available'

        Phase = str(find_between_r(tables,"Study Phase","Study Design"))
        Phase = Phase.strip()
        
        Summary = (str(find_between_r(tables,"Brief Summary","Detailed")))
        #StartDate.append((str(find_between_r(tables,"Start Date","Estimated Primary"))))
        #CompletionDate.append((str(find_between_r(tables,"Primary Completion Date","Current Primary"))))
        DetailedDescription = ((str(find_between_r(tables,"Detailed Description","Study Phase")).split("Study Type")[0]))
        #Condition is blank
        Ages = str(find_between_r(tables,"Ages","Accepts Healthy Volunteers"))
        Identifier = str(url.split("record/")[1].split("?")[0])
        Criteria = str(find_between_r(tables,"Eligibility Criteria","Sex/Gender"))
        Drug = str(differences[h])
        
        Input_Data['B' + str(ResultsRowCounter)] = Drug
        Input_Data['C' + str(ResultsRowCounter)] = ConditionWebsite
        Input_Data['D' + str(ResultsRowCounter)] = Phase
        Input_Data['F' + str(ResultsRowCounter)] = StartDate
        Input_Data['G' + str(ResultsRowCounter)] = CompletionDate
        Input_Data['H' + str(ResultsRowCounter)] = Summary
        Input_Data['I' + str(ResultsRowCounter)] = DetailedDescription            
        Input_Data['J' + str(ResultsRowCounter)] = Criteria
        Input_Data['K' + str(ResultsRowCounter)] = Ages          
        
        CompOne = StartDate
        CompTwo = CompletionDate
        Months = [CompOne, CompTwo]
        Years = [CompOne, CompTwo]

        m = 0
        while(m < 2):
            if(Months[m][0:3]=='Jan'):
                Months[m] = 1
            elif(Months[m][0:3]=='Feb'):
                Months[m] = 2
            elif(Months[m][0:3]=='Mar'):
                Months[m] = 3
            elif(Months[m][0:3]=='Apr'):
                Months[m] = 4
            elif(Months[m][0:3]=='May'):
                Months[m] = 5
            elif(Months[m][0:3]=='Jun'):
                Months[m] = 6
            elif(Months[m][0:3]=='Jul'):
                Months[m] = 7
            elif(Months[m][0:3]=='Aug'):
                Months[m] = 8
            elif(Months[m][0:3]=='Sep'):
                Months[m] = 9
            elif(Months[m][0:3]=='Oct'):
                Months[m] = 10
            elif(Months[m][0:3]=='Nov'):
                Months[m] = 11
            elif(Months[m][0:3]=='Dec'):
                Months[m] = 12
            else:
                Months[m] = 0
            m += 1
    
        n = 0
        while(n < 2):
            if(Years[n] == '0'):
                pass
            else:
                Years[n] = Years[n][-4:]
            n += 1
    
        LengthTrial = 0
        
        if(Years[0].isdigit() and Years[1].isdigit()):
            if(int(Years[1]) - int(Years[0]) > 1):
                LengthTrial = ((12 - Months[0]) + Months[1]) + 12*(int(Years[1]) - int(Years[0]) - 1)
            elif(int(Years[1]) - int(Years[0]) == 1):
                LengthTrial = ((12 - Months[0]) + Months[1])
            elif(int(Years[1]) - int(Years[0]) == 0):
                LengthTrial = int(Months[1]) - int(Months[0])
            else:
                LengthTrial = 0
            if(LengthTrial < 0):
                LengthTrial = abs(LengthTrial)
        else:
            LengthTrial = 'Not Available'
        
        Input_Data['E' + str(ResultsRowCounter)] = LengthTrial
        
        ResultsRowCounter += 1
        l += 1
    
    print("Finished Item #" + str(i) + " of " + str(len(differences))) + ". Rows of data Filled: " + str(ResultsRowCounter) + "."

    i += 1
    h += 1

    #Time delay, to help prevent connection issues 
    firstBotTrickSec = random.random() * random.randint(5, 30)
    firstBotTrick = int(firstBotTrickSec)
    time.sleep(firstBotTrick)

wb.save(open_file)
print("Done.")