#-------------------------------------------------------------------------------
# Name:        Excel MultiLine TimeSaver
# Purpose:     Takes input from an excel file, searches Clinicaltrials.gov for
#              information, then extracts it and records in a separate sheet.
#
# Author:      Jonathan Thomas

# Licence:     MIT "Free Beer" License
#-------------------------------------------------------------------------------

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
import StringIO
import re
import unicodedata

import Tkinter
import tkFileDialog
import html5lib
import time
import random


#remember to have requests, BeautifulSoup4, requests[security], and in your packages for this to work!
import requests
from bs4 import BeautifulSoup
#Define starting constants:

MainCounter = 2
#Beginning and Ending Cells in your worksheet to Process
StartCell = 632
EndCell = 663

#Code Shamelessly copied from StackExchange user HennyH
def search(text,n):
    '''Searches for text, and retrieves n words either side of the text, which are returned seperately'''
    word = r"\W*([\w]+)"
    groups = re.search(r'{}\W*{}{}'.format(word*n,'place',word*n), text).groups()
    return groups[:n],groups[n:]

def find_all(a_str, sub):
    start = 0
    while True:
        start = a_str.find(sub, start)
        if start == -1: return
        yield start
        start += len(sub) # use start += 1 to find overlapping matches


def find_between_r( s, first, last ):
    try:
        start = s.rindex( first ) + len( first )
        end = s.rindex( last, start )
        return s[start:end]
    except ValueError:
        return ""

#Open the file in question:

open_file = tkFileDialog.askopenfilename ()

#Dialog Box opening complete

print("Now opening: " + str(open_file))

#Performing initial IO on workbooks
wb = load_workbook(filename = open_file)
blank = Workbook()
ws = wb.active

sheet_ranges = wb["Drugs"]
ResultsInput = wb["Results"]
CellIter = StartCell
FinishedLinesTest = []
firstBotTrick = 0
firstBotTrickSec = 0
while(CellIter<EndCell):

    OBJ = sheet_ranges['A' + str(CellIter)].value

    RawLines = StringIO.StringIO(str(OBJ))


    #Read the Excel file and put into array
    ActualLines = RawLines.readlines()

    FinishedLines = [x.strip() for x in ActualLines]
    
    i = 0
    StringToAdd=[]
    while(i<len(FinishedLines)):
        if((FinishedLines[i].find(",")) != -1):
            StringToAdd.append(FinishedLines[i].split(",")[0])
        i = i + 1

    FinishedLines = FinishedLines + StringToAdd
    #print(FinishedLines)
    #print(StringToAdd)

    if(FinishedLinesTest == FinishedLines):
        CellIter = CellIter + 1
        continue

    #Grab info from web pages search terms

    ResultsNumber = []
    kame = 0
    kameTestList = []
    secondBotTrick = 0
    secondBotTrickSec = 0
    while(kame<len(FinishedLines)):

        if(FinishedLines[kame] in kameTestList):
            kame = kame + 1
            continue

        secondBotTrickSec = random.random() * random.choice([10, 20, 30])
        secondBotTrick = int(secondBotTrickSec)
        time.sleep(secondBotTrick)

        #Find Queries and count number
        url = 'https://clinicaltrials.gov/ct2/results?cond=&term=' + FinishedLines[kame].replace(' ','+') + '&cntry=&state=&city=&dist='
        r=requests.get(url)
        html = r.text
        soup = BeautifulSoup(r.text,"html5lib")

        #print(url)

        WebLines =  StringIO.StringIO(str(soup))
        FinWebLines = WebLines.readlines()

        i = 0
        DataLine = 0
        while(i<len(FinWebLines)):
            if(FinWebLines[i].find("found for:") != -1):
                #print(FinWebLines[i])
                DataLine = i
                break
            i = i + 1

        DataLine = FinWebLines[i].split()
        i = 0
        a = 0
        while(i <len(DataLine)):
            if(DataLine[i].find("found") != -1):
                a = i-2
                break
            i = i + 1

        ResultsNumber = DataLine[a].split(">")[-1]
        if(ResultsNumber == "No"):
            ResultsNumber = 0

        ResultsNumber = int(ResultsNumber)

        #Get links to search results

        links = []

        for item in soup.find_all('a'):
            links.append((item.get('href')))

        links = [str(links[x]) for x in range(len(links))]
        #print(len(links))
        glinks = []
        i = 0
        while(i<len(links)):
            if((links[i].find("/ct2/show/") != -1 )):
                glinks.append("https://clinicaltrials.gov" + links[i])
            i = i + 1

        links = glinks
        #print(links)

        #Go into each link and check for criterion

        i = 0
        GoodStudies = []
        INCLUDE = 1
        while(i<len(links)):
            url = links[i]
            r=requests.get(url)
            html = r.text
            soup = BeautifulSoup(html,"html5lib")
            [s.extract() for s in soup(['style', 'script', '[document]', 'head', 'title'])]
            SearchText = soup.getText()
            SearchText = SearchText.encode('utf-8')
            #First Set of Criterion
            if(SearchText.find("Suspended") == -1 and SearchText.find("Unknown") == -1 and SearchText.find("Terminated") == -1):
                #Second Set of Criterion
                if(SearchText.find("Observational") == -1 and SearchText.find("Expanded Access") == -1):
                    #Fourth Set of Criterion
                    if(SearchText.find("Phase 4") == -1):
                        GoodStudies.append(links[i])

            i = i + 1
            print("Study " + str(i) + " out of: " + str(len(links)))

        #print(GoodStudies)
        i = 0
        j = 0
        link = []
        Condition = []
        Ages = []
        Identifier = []
        Criteria = []
        Summary = []
        TimeofStudy = []
        DetailedDescription = []
        StartDate = []
        CompletionDate = []
        thirdBotTrick = 0
        thirdBotTrickSec = 0
        while(i<len(GoodStudies)):

            thirdBotTrickSec = random.random() * random.choice([20, 30, 40, 50, 60])
            thirdBotTrick = int(thirdBotTrickSec)
            time.sleep(thirdBotTrick)

            url = GoodStudies[i].replace("show/","show/record/")
            r=requests.get(url)
            html = r.text
            soup = BeautifulSoup(html,"html5lib")

            tables = []
            [s.extract() for s in soup(['style', 'script', '[document]', 'head', 'title'])]
            tables = soup.getText()
            tables = tables.encode('utf-8')
            Time = find_between_r(tables,"Last Update Posted Date","Start")
            Time = Time.split()
            if(Time == []):
                Time.append(2000)
            TimeofStudy.append(Time[-1])

            if(TimeofStudy[-1]<2012):
                i += 1
                continue


            link.append(url)
            mina = list(find_all(tables,"Condition"))
            minb = list(find_all(tables,"Intervention"))
            if(minb == []):
                minb.append(0)
            if(mina == []):
                mina.append(0)

            checker = 0
            while(checker<len(minb)):
                if(minb[checker]<=min(mina) or minb[checker]>=25000):
                    minb[checker] = 100000000
                checker = checker + 1
            smallcont = []
            storagecont = []
            smallstorage = []
            u = 0
            p = 0
            while(u<len(mina)):
                p = 0
                while((p<len(minb))):
                    if(minb[p] > mina[u]):
                        storagecont.append(abs(mina[u] -  minb[p]))
                    else:
                        storagecont.append(10000000)
                    p = p + 1
                smallstorage.append(storagecont.index(min(storagecont)))
                smallcont.append(storagecont[storagecont.index(min(storagecont))])
                storagecont = []
                u = u + 1

            PosOne = mina[smallcont.index(min(smallcont))]
            PosTwo = minb[smallstorage[smallcont.index(min(smallcont))]]

            Condition.append(str(tables[PosOne:PosTwo]).replace("Condition","").replace("ICMJE",""))

            mina = list(find_all(tables,"Start"))
            if(mina == []):
                mina.append(0)
            minb = list(find_all(tables,"Primary"))
            if(minb == []):
                minb.append(0)


            #print(mina)
            #print(minb)
            smallcont = []
            storagecont = []
            smallstorage = []
            u = 0
            p = 0
            while(u<len(mina)):
                p = 0
                while((p<len(minb))):
                    storagecont.append(abs(mina[u] -  minb[p]))
                    p = p + 1
                smallstorage.append(storagecont.index(min(storagecont)))
                smallcont.append(storagecont[storagecont.index(min(storagecont))])
                storagecont = []
                u = u + 1
            PosOne = mina[smallcont.index(min(smallcont))]
            PosTwo = minb[smallstorage[smallcont.index(min(smallcont))]]
            StartDate.append(str(tables[PosOne:PosTwo]).replace("Start","").replace("Date","").replace("ICMJE","").replace("Estimated",""))

            mina = list(find_all(tables,"Completion Date"))
            if(mina == []):
                mina.append(0)
            minb = list(find_all(tables,"Primary"))
            if(minb == []):
                minb.append(0)

            checker = 0
            while(checker<len(minb)):
                if(minb[checker]<=min(mina) or minb[checker]>=8000):
                    minb[checker] = 100000000
                checker = checker + 1
            #print(mina)
            #print(minb)
            smallcont = []
            storagecont = []
            smallstorage = []
            u = 0
            p = 0
            while(u<len(mina)):
                p = 0
                while((p<len(minb))):
                    storagecont.append(abs(mina[u] -  minb[p]))
                    p = p + 1
                smallstorage.append(storagecont.index(min(storagecont)))
                smallcont.append(storagecont[storagecont.index(min(storagecont))])
                storagecont = []
                u = u + 1
            PosOne = mina[smallcont.index(min(smallcont))]
            PosTwo = minb[smallstorage[smallcont.index(min(smallcont))]]
            CompletionDate.append(str(tables[PosOne:PosTwo]).replace("Primary","").replace("Date","").replace("Current","").replace("Estimated","").replace("Completion",""))

            Summary.append(str(find_between_r(tables,"Brief Summary","Detailed")))
            #StartDate.append((str(find_between_r(tables,"Start Date","Estimated Primary"))))
            #CompletionDate.append((str(find_between_r(tables,"Primary Completion Date","Current Primary"))))
            DetailedDescription.append((str(find_between_r(tables,"Detailed Description","Study Phase")).split("Study Type")[0]))
            #Condition is blank
            Ages.append(str(find_between_r(tables,"Ages","Accepts Healthy Volunteers")))
            Identifier.append(str(url.split("record/")[1].split("?")[0]) )
            Criteria.append(str(find_between_r(tables,"Eligibility Criteria","Sex/Gender")))
            Drug = str(FinishedLines[kame])
            ResultsInput["A" + str(MainCounter)] = Drug
            ResultsInput["B" + str(MainCounter)] = Summary[j]
            ResultsInput["C" + str(MainCounter)] = DetailedDescription[j]
            ResultsInput["D" + str(MainCounter)] = StartDate[j]
            ResultsInput["E" + str(MainCounter)] = CompletionDate[j]
            ResultsInput["H" + str(MainCounter)] = Criteria[j]
            ResultsInput["G" + str(MainCounter)] = Ages[j]
            ResultsInput["I" + str(MainCounter)] = Identifier[j]
            ResultsInput["J" + str(MainCounter)] = link[j]
            ResultsInput["K" + str(MainCounter)] = Condition[j]
            i = i + 1
            j += 1
            MainCounter = MainCounter + 1
            #print(Condition)
        kameTestList.append(FinishedLines[kame])
        kame = kame + 1



        wb.save(open_file)
    #print(Identifier)
    #print(Criteria)
    #print(Ages)
    #print(Condition)
    #print(Summary)
    #print(url)
    FinishedLinesTest = FinishedLines
    print("Finished Scanning Cell: " + str(CellIter) + " of " + str(EndCell)) + ". Rows of data Filled: " + str(MainCounter) + "."
    CellIter = CellIter + 1

    firstBotTrickSec = random.random() * random.randint(0, 15)
    firstBotTrick = int(firstBotTrickSec)
    time.sleep(firstBotTrick)


print("Done")







